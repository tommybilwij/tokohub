"""CSV and Excel file parser for receipt import."""

import csv
import io
import logging

import openpyxl

logger = logging.getLogger(__name__)

# Column name aliases for auto-detection
_NAME_ALIASES = {'nama_barang', 'item_name', 'artname', 'nama', 'name', 'barang', 'item', 'product'}
_QTY_ALIASES = {'qty', 'jumlah', 'quantity', 'jlh', 'kuantitas', 'pcs'}
_PRICE_ALIASES = {'harga', 'price', 'harga_beli', 'unit_price', 'hbeli', 'cost'}


def _detect_columns(headers):
    """Map header names to our standard fields: name, qty, price."""
    mapping = {}
    normalized = [h.lower().strip().replace(' ', '_') for h in headers]

    for i, col in enumerate(normalized):
        if col in _NAME_ALIASES and 'name' not in mapping:
            mapping['name'] = i
        elif col in _QTY_ALIASES and 'qty' not in mapping:
            mapping['qty'] = i
        elif col in _PRICE_ALIASES and 'price' not in mapping:
            mapping['price'] = i

    return mapping


def _parse_number(value):
    """Parse a number from string, handling commas and dots."""
    if value is None:
        return 0
    s = str(value).strip().replace(',', '')
    try:
        return float(s)
    except ValueError:
        return 0


def parse_csv(file_path=None, file_stream=None):
    """Parse a CSV file and return structured items.

    Returns list of dicts: [{name, qty, price}]
    """
    if file_stream:
        text = file_stream.read()
        if isinstance(text, bytes):
            text = text.decode('utf-8', errors='replace')
        reader = csv.reader(io.StringIO(text))
    else:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            return _process_csv_reader(reader)

    return _process_csv_reader(reader)


def _process_csv_reader(reader):
    """Process CSV reader rows into item list."""
    rows = list(reader)
    if not rows:
        return []

    mapping = _detect_columns(rows[0])
    if 'name' not in mapping:
        # Fallback: assume first column is name
        mapping = {'name': 0, 'qty': 1, 'price': 2}
        data_rows = rows  # No header row detected
    else:
        data_rows = rows[1:]

    items = []
    for row in data_rows:
        if not row or not row[mapping['name']].strip():
            continue
        items.append({
            'name': row[mapping['name']].strip(),
            'qty': _parse_number(row[mapping.get('qty', 1)] if mapping.get('qty') is not None and mapping['qty'] < len(row) else 1),
            'price': _parse_number(row[mapping.get('price', 2)] if mapping.get('price') is not None and mapping['price'] < len(row) else 0),
        })

    logger.info("CSV parsed: %d items", len(items))
    return items


def parse_excel(file_path=None, file_stream=None):
    """Parse an Excel (.xlsx) file and return structured items.

    Returns list of dicts: [{name, qty, price}]
    """
    if file_stream:
        wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(c or '') for c in rows[0]]
    mapping = _detect_columns(headers)

    if 'name' not in mapping:
        mapping = {'name': 0, 'qty': 1, 'price': 2}
        data_rows = rows
    else:
        data_rows = rows[1:]

    items = []
    for row in data_rows:
        name_val = str(row[mapping['name']] or '').strip() if mapping['name'] < len(row) else ''
        if not name_val:
            continue
        items.append({
            'name': name_val,
            'qty': _parse_number(row[mapping.get('qty', 1)] if mapping.get('qty') is not None and mapping['qty'] < len(row) else 1),
            'price': _parse_number(row[mapping.get('price', 2)] if mapping.get('price') is not None and mapping['price'] < len(row) else 0),
        })

    logger.info("Excel parsed: %d items", len(items))
    return items
